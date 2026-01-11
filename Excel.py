import pandas as pd
from datetime import datetime

from Joint_zed import Joint
import openpyxl
import Settings as s
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import os


PATIENTS_FILE = "Patients.xlsx"


def ensure_patients_file_exists():
    """
    Creates the Patients.xlsx file with required sheets if it doesn't exist.
    This ensures the system works even on first run.
    Structure matches the existing Patients.xlsx template exactly.
    """
    if os.path.exists(PATIENTS_FILE):
        return True  # File already exists
    
    print(f"Creating {PATIENTS_FILE}...")
    
    try:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create patients_details sheet (exact match to template)
        details_sheet = workbook.create_sheet("patients_details")
        details_headers = [
            "ID", "gender", "number of exercises", 
            "number of repetitions in each exercise", "rate"
        ]
        for col, header in enumerate(details_headers, 1):
            details_sheet.cell(row=1, column=col, value=header)
        
        # Create patients_history_of_trainings sheet
        history_sheet = workbook.create_sheet("patients_history_of_trainings")
        history_sheet.cell(row=1, column=1, value="ID")
        
        # Create patients_exercises sheet (exact match to template)
        exercises_sheet = workbook.create_sheet("patients_exercises")
        exercises_headers = [
            "ID",
            "ball_bend_elbows",
            "ball_raise_arms_above_head",
            "ball_switch",
            "ball_open_arms_and_forward",
            # NOTE: ball_open_arms_above_head removed - no UI checkbox and no Gymmy implementation
            "band_open_arms",
            "band_open_arms_and_up",
            "band_up_and_lean",
            "band_straighten_left_arm_elbows_bend_to_sides",
            "band_straighten_right_arm_elbows_bend_to_sides",
            "stick_bend_elbows",
            "stick_bend_elbows_and_up",
            "stick_raise_arms_above_head",
            "stick_switch",
            "stick_up_and_lean",
            "weights_open_arms_and_forward",
            "weights_abduction",
            "notool_hands_behind_and_lean",
            "notool_right_hand_up_and_bend",
            "notool_left_hand_up_and_bend",
            "notool_raising_hands_diagonally",
            "notool_right_bend_left_up_from_side",
            "notool_left_bend_right_up_from_side"
        ]
        for col, header in enumerate(exercises_headers, 1):
            exercises_sheet.cell(row=1, column=col, value=header)
        
        # Save the workbook
        workbook.save(PATIENTS_FILE)
        print(f"{PATIENTS_FILE} created successfully with all required sheets!")
        return True
        
    except Exception as e:
        print(f"Error creating {PATIENTS_FILE}: {e}")
        return False


def create_and_open_folder(folder_path):
    try:
        # Create the folder if it doesn't exist
        os.makedirs(folder_path, exist_ok=True)
        print(f"Directory created or already exists: {folder_path}")

        # Check if the folder exists after creation
        if not os.path.exists(folder_path):
            raise FileNotFoundError(f"Directory still not found: {folder_path}")

    except Exception as e:
        print(f"An error occurred: {e}")


# #creats a new workbook to each training
# def create_workbook_for_training():
#     datetime_string = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
#     workbook_name = f"Patients/{s.chosen_patient_ID}/{datetime_string}.xlsx"
#     s.training_workbook_path = workbook_name
#     s.training_workbook = openpyxl.Workbook()  # Do not pass the filename here
#     s.training_workbook.save(s.training_workbook_path)  # Save the workbook after creating it


def create_workbook_for_training():
    # Use ISO format for better sorting: YYYY-MM-DD_HH-MM-SS
    datetime_string = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    # Determine folder based on session type
    if getattr(s, 'is_rom_assessment_mode', False):
        # ROM Assessment - save in ROM_Assessments/datetime/ folder
        session_folder = f"Patients/{s.chosen_patient_ID}/ROM_Assessments/{datetime_string}"
        os.makedirs(session_folder, exist_ok=True)
        workbook_name = f"{session_folder}/Calibration_Data.xlsx"
        print(f"[Excel] Creating ROM session folder: {session_folder}")
    else:
        # Regular Training - save in Trainings/datetime/ folder
        session_folder = f"Patients/{s.chosen_patient_ID}/Trainings/{datetime_string}"
        os.makedirs(session_folder, exist_ok=True)
        workbook_name = f"{session_folder}/Session_Data.xlsx"
        print(f"[Excel] Creating Training session folder: {session_folder}")
    
    # Store session folder path for use by graph/table functions
    s.current_session_folder = session_folder
    
    s.training_workbook_path = workbook_name
    s.training_workbook = openpyxl.Workbook()
    s.training_workbook.save(s.training_workbook_path)



#returns a specific value by ID and name of the column
def find_value_by_colName_and_userID(workbook_path, worksheet_name, ID, target_col):
    try:
        # Load the workbook
        workbook = load_workbook(workbook_path)

        # Access the worksheet
        worksheet = workbook[worksheet_name]
        target_row=None

        # Search for the target value in the first column
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1), start=2):
            for cell in row:
                if str(cell.value) == ID:
                    target_row= row_number  # Return row number and value from specified column
                    break
            if target_row is not None:
                break

        for col in worksheet.iter_cols():
            # Check if the first cell in the column matches the specified column name
            if col[0].value == target_col:
                # Return the entire column
                return col[target_row-1].value

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


#returns the value of success in a specific training and specific exercise
def get_success_number(file_path, exercise):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Check if the worksheet exists
        if "success_worksheet" not in workbook.sheetnames:
            print(f"Worksheet success not found in the workbook.")
            return None

        # Select the worksheet
        worksheet = workbook["success_worksheet"]

        # Iterate through the rows in the first column and search for the value
        for row in worksheet.iter_rows(values_only=True):
            if row[0] == exercise:
                # Return the value from the second column
                return row[1]

        # If the value is not found, return None
        return None

    except FileNotFoundError:
        print(f"File success not found.")
        return None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# #returns the value of effort rate in a specific training and specific exercise
# def get_effort_number(file_path, exercise):
#     try:
#         # Load the workbook
#         workbook = openpyxl.load_workbook(file_path)
#
#         # Check if the worksheet exists
#         if "success_and_effort" not in workbook.sheetnames:
#             print(f"Worksheet success not found in the workbook.")
#             return None
#
#         # Select the worksheet
#         worksheet = workbook["success_and_effort"]
#
#         return worksheet[1,2]
#
#
#     except FileNotFoundError:
#         print(f"File success not found.")
#         return None
#

def wf_joints(ex_name, list_joints):
    # Check if "Sheet" exists and delete it
    if "Sheet" in s.training_workbook.sheetnames:
        sheet_to_delete = s.training_workbook["Sheet"]
        s.training_workbook.remove(sheet_to_delete)

    # Create a new sheet with the provided name
    worksheet1 = s.training_workbook.create_sheet(ex_name[:31])
    col = 0

    for l in range(0, len(list_joints)):
        worksheet1.cell(row=1, column=col + 1, value=col + 1)

        row = 2
        for j in list_joints[l]:
            if isinstance(j, Joint):  # Check if j is a Joint object
                j_ar = j.joint_to_array()
                for i in range(len(j_ar)):
                    worksheet1.cell(row=row, column=col + 1, value=str(j_ar[i]))
                    row += 1
            else:
                # Handle other types appropriately, e.g., just write the value to the worksheet
                worksheet1.cell(row=row, column=col + 1, value=str(j))
                row += 1

        col += 1

    # Save the workbook
    success_worksheet()
    s.training_workbook.save(s.training_workbook_path)
    create_graphs_and_tables(ex_name, list_joints)



def create_graphs_and_tables(exercise, list_joints):
    try:
        if get_number_of_angles_in_exercise(exercise) == 1:
            one_angle_graph_and_table(exercise, list_joints)
        if get_number_of_angles_in_exercise(exercise) == 2:
            two_angles_graph_and_table(exercise, list_joints)
        if get_number_of_angles_in_exercise(exercise) == 3:
            three_angles_graph_and_table(exercise, list_joints)



    except (pd.errors.ParserError, FileNotFoundError):
        # Handle the case where the sheet is not found
        pass  # Continue to the next iteration
    except ValueError as ve:
        # Handle other specific errors
        pass  # Continue to the next iteration

def get_name_by_exercise(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'name']
    # Return the exact value, including formatting (e.g., RTL or LTR)
    return result.iloc[0] if not result.empty else None


def get_number_of_angles_in_exercise(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'number of angles']
    return result.iloc[0] if not result.empty else None


def get_equipment(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'equipment']
    return result.iloc[0] if not result.empty else None


def get_repetitions_per_count(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'counts after one or two repetitions']
    return result.iloc[0] if not result.empty else None


def get_files_names_by_start_word(word):
    """
    Retrieves all file names in the directory that start with 'dont_recognize'.
    Returns:
    - List[str]: A list of matching file names without extensions.
    """
    matching_file_names = []

    if not os.path.exists(s.audio_path):
        print(f"Directory does not exist: {s.audio_path}")
        return matching_file_names

    for file_name in os.listdir(s.audio_path):
        if not file_name.startswith(word):
            continue
        name_without_extension, _ = os.path.splitext(file_name)
        matching_file_names.append(name_without_extension)

    return matching_file_names

def one_angle_graph_and_table(exercise_name, list_joints):
    if (list_joints!=[]):
        last_two_values = [entry[-2:] for entry in list_joints] #extract from each record the last 2 values (the angles)
        right_angles = [sublist[0] for sublist in last_two_values] #the right angle from each record
        left_angles = [sublist[1] for sublist in last_two_values] #the left angle from each record


        #extract the joints names and create graphs names
        first_values= list_joints[0]
        first_6_values= first_values[:6]
        joints_names = [str(sample).split()[0] for sample in first_6_values]
        first_graph_name= joints_names[0]+", "+joints_names[1]+", "+joints_names[2]+" 1"
        second_graph_name= joints_names[3]+", "+joints_names[4]+", "+joints_names[5]+" 2"

        #create a list of x values
        length= len(list_joints)
        measurement_num = list(range(1, length + 1))

        #create a data dic for graph
        data = {
        first_graph_name: {'x': measurement_num, 'y': right_angles},
        second_graph_name: {'x': measurement_num, 'y': left_angles}}

        create_and_save_graph(data, exercise_name)
        create_and_save_table_with_calculations(data, exercise_name)


def two_angles_graph_and_table(exercise_name, list_joints):
    if (list_joints!=[]):
        last_four_values = [entry[-4:] for entry in list_joints]  # extract from each record the last 4 values (the angles)
        right_angles = [sublist[0] for sublist in last_four_values]  # the right angle from each record
        left_angles = [sublist[1] for sublist in last_four_values]  # the left angle from each record
        right_angles2 = [sublist[2] for sublist in last_four_values]  # the second right angle from each record
        left_angles2 = [sublist[3] for sublist in last_four_values]  # the second left angle from each record

        # extract the joints names and create graphs names
        first_values = list_joints[0]
        first_12_values = first_values[:12]
        joints_names = [str(sample).split()[0] for sample in first_12_values]
        first_graph_name = joints_names[0] + ", " + joints_names[1] + ", " + joints_names[2]+" 1"
        second_graph_name = joints_names[3] + ", " + joints_names[4] + ", " + joints_names[5]+" 2"
        third_graph_name = joints_names[6] + ", " + joints_names[7] + ", " + joints_names[8]+" 3"
        fourth_graph_name = joints_names[9] + ", " + joints_names[10] + ", " + joints_names[11]+" 4"

        # create a list of x values
        length = len(list_joints)
        measurement_num = list(range(1, length + 1))

        # create a data dic for graph
        data = {
            first_graph_name: {'x': measurement_num, 'y': right_angles},
            second_graph_name: {'x': measurement_num, 'y': left_angles},
            third_graph_name: {'x': measurement_num, 'y': right_angles2},
            fourth_graph_name: {'x': measurement_num, 'y': left_angles2}
        }

        create_and_save_graph(data, exercise_name)
        create_and_save_table_with_calculations(data, exercise_name)


def three_angles_graph_and_table(exercise_name, list_joints):
    if (list_joints!=[]):
        last_four_values = [entry[-6:] for entry in list_joints]  # extract from each record the last 6 values (the angles)
        right_angles = [sublist[0] for sublist in last_four_values]  # the right angle from each record
        left_angles = [sublist[1] for sublist in last_four_values]  # the left angle from each record
        right_angles2 = [sublist[2] for sublist in last_four_values]  # the second right angle from each record
        left_angles2 = [sublist[3] for sublist in last_four_values]  # the second left angle from each record
        right_angles3 = [sublist[4] for sublist in last_four_values]  # the third right angle from each record
        left_angles3 = [sublist[5] for sublist in last_four_values]  # the third left angle from each record

        # extract the joints names and create graphs names
        first_values = list_joints[0]
        first_18_values = first_values[:18]
        joints_names = [str(sample).split()[0] for sample in first_18_values]
        first_graph_name = joints_names[0] + ", " + joints_names[1] + ", " + joints_names[2]+" 1"
        second_graph_name = joints_names[3] + ", " + joints_names[4] + ", " + joints_names[5]+" 2"
        third_graph_name = joints_names[6] + ", " + joints_names[7] + ", " + joints_names[8]+" 3"
        fourth_graph_name = joints_names[9] + ", " + joints_names[10] + ", " + joints_names[11]+" 4"
        fifth_graph_name = joints_names[12] + ", " + joints_names[13] + ", " + joints_names[14]+" 5"
        sixth_graph_name = joints_names[15] + ", " + joints_names[16] + ", " + joints_names[17]+" 6"

        # create a list of x values
        length = len(list_joints)
        measurement_num = list(range(1, length + 1))

        # create a data dic for graph
        data = {
            first_graph_name: {'x': measurement_num, 'y': right_angles},
            second_graph_name: {'x': measurement_num, 'y': left_angles},
            third_graph_name: {'x': measurement_num, 'y': right_angles2},
            fourth_graph_name: {'x': measurement_num, 'y': left_angles2},
            fifth_graph_name: {'x': measurement_num, 'y': right_angles3},
            sixth_graph_name: {'x': measurement_num, 'y': left_angles3}
        }

        create_and_save_graph(data, exercise_name)
        create_and_save_table_with_calculations(data, exercise_name)


# def two_joints_distance_graphs_and_table(exercise_name, list_joints):
#     if (list_joints!=[]):
#         last_two_values = [entry[-2:] for entry in list_joints] #extract from each record the last 2 values (the angles)
#         right_distance = [sublist[0] for sublist in last_two_values] #the right angle from each record
#         left_distance = [sublist[1] for sublist in last_two_values] #the left angle from each record
#
#
#         #extract the joints names and create graphs names
#         first_values= list_joints[0]
#         first_4_values= first_values[:4]
#         joints_names = [str(sample).split()[0] for sample in first_4_values] #takes from each joint only the first string which is the name
#         first_graph_name= "Distance " + joints_names[0] + " To " + joints_names[1]
#         second_graph_name= "Distance " + joints_names[2] + " To " + joints_names[3]
#
#         #create a list of x values
#         length= len(list_joints)
#         measurement_num = list(range(1, length + 1))
#
#         #create a data dic for graph
#         data = {
#         first_graph_name: {'x': measurement_num, 'y': right_distance},
#         second_graph_name: {'x': measurement_num, 'y': left_distance}}
#
#         create_and_save_graph(data, exercise_name)
#         create_and_save_table_with_calculations(data, exercise_name)


def create_and_save_graph(data, exercise):
    # Get session folder (created by create_workbook_for_training)
    session_folder = getattr(s, 'current_session_folder', None)
    
    if session_folder is None:
        # Fallback to old behavior if session_folder not set
        timestamp = s.starts_and_ends_of_stops[0]
        start_dt = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d_%H-%M-%S")
        session_folder = f"Patients/{s.chosen_patient_ID}/Trainings/{start_dt}"
    
    # Create exercise folder inside session folder (only when needed!)
    exercise_folder = f"{session_folder}/{exercise}"
    os.makedirs(exercise_folder, exist_ok=True)
    
    # Iterate over each plot data
    for plot_name, plot_data in data.items():
        # Create a new plot
        y_series = pd.Series(plot_data['y'])
        x_values = plot_data['x']
        y_values = y_series.values  # Keep NaN values

        # Check if all values in y_series are NaN or empty
        if y_series.isnull().all() or y_series.count() < 10:
            # Save a "null graph"
            plot_filename = f'{exercise_folder}/{plot_name}.jpeg'

            # Create a "null graph"
            plt.figure()
            plt.text(
                0.5, 0.5, "No Data Available", fontsize=18, color="gray", ha="center", va="center", alpha=0.7
            )
            plt.axis("off")  # Hide axes
            plt.title(plot_name[:-2], fontsize=16, weight="bold", y=0.9)
            plt.savefig(plot_filename, bbox_inches="tight", pad_inches=0, dpi=100)
            plt.close()
            continue

        # Plot the graph; matplotlib handles NaN by breaking the line
        plt.plot(x_values, y_values)

        # Highlight NaN values with red dots at y=0
        nan_indices = np.where(pd.isnull(y_values))[0]  # Find indices of NaN values
        if len(nan_indices) > 0:
            plt.scatter(
                [x_values[i] for i in nan_indices],
                [0 for _ in nan_indices],  # Placeholders at y=0 for NaN
                color='red',
                label="No Data",
                zorder=5
            )

        # Set the font size
        fontsize = 16

        plt.xlabel('מספר מדידה'[::-1], fontsize=fontsize, weight='bold')
        plt.ylabel('זווית'[::-1], fontsize=fontsize, weight='bold')
        plt.title(plot_name[:-2], fontsize=16, weight="bold", y=1)

        # Save the plot as an image file
        plot_filename = f'{exercise_folder}/{plot_name}.jpeg'
        plt.savefig(plot_filename, dpi=100)
        plt.close()  # Close the plot to clear the figure
    
    print(f"[Excel] Saved graphs to: {exercise_folder}")


def success_worksheet():
    exercise, success_count = list(s.ex_list.items())[-1]

    # Check if the sheet exists
    if "success_worksheet" in s.training_workbook.sheetnames:
        # Get the existing worksheet
        success_sheet = s.training_workbook["success_worksheet"]
    else:
        # Create a new sheet if it doesn't exist
        success_sheet = s.training_workbook.create_sheet("success_worksheet")
        # Write headers in the first row (row=1)
        success_sheet.cell(row=1, column=1, value="exercise")
        success_sheet.cell(row=1, column=2, value="number of successful repetitions")

    # Find the first empty row in the worksheet
    row = success_sheet.max_row + 1

    # Write the exercise name and success count into the worksheet
    success_sheet.cell(row=row, column=1, value=exercise)
    success_sheet.cell(row=row, column=2, value=success_count)

    # Save the workbook after writing to the sheet
    s.training_workbook.save(s.training_workbook_path)


def find_and_change_values_exercises(new_values_dict, headers_row=1):
    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_exercises"]

    # Find the column indices based on the header names
    column_indices = {}
    for header_name, new_value in new_values_dict.items():
        for cell in sheet[headers_row]:
            if cell.value == header_name:
                column_indices[header_name] = cell.column
                break

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if str(cell.value) == s.chosen_patient_ID:
            # Update the values in the corresponding columns
            for header_name, column_index in column_indices.items():
                sheet.cell(row=cell.row, column=column_index, value=new_values_dict[header_name])

    # Save the changes
    workbook.save(file_path)


def calculate_training_length():
    if len(s.starts_and_ends_of_stops) % 2 != 0:
        s.starts_and_ends_of_stops.pop(-2)

    training_length = 0.0  # total in seconds

    for i in range(0, len(s.starts_and_ends_of_stops), 2):
        start_time = s.starts_and_ends_of_stops[i]
        end_time = s.starts_and_ends_of_stops[i + 1]
        print(f"Start: {start_time}, End: {end_time}")
        training_length += (end_time - start_time)

    training_length= training_length/60.0
    return training_length # float seconds


def find_and_change_values_patients(new_values_dict, headers_row=1):
    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_details"]

    # Find the column indices based on the header names
    column_indices = {}
    for header_name, new_value in new_values_dict.items():
        for cell in sheet[headers_row]:
            if cell.value == header_name:
                column_indices[header_name] = cell.column
                break

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if str(cell.value) == s.chosen_patient_ID:
            # Update the values in the corresponding columns
            for header_name, column_index in column_indices.items():
                sheet.cell(row=cell.row, column=column_index, value=new_values_dict[header_name])

    # Save the changes
    workbook.save(file_path)



def find_and_add_training_to_patient(headers_row=1):
    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_history_of_trainings"]

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if str(cell.value) == s.chosen_patient_ID:
            # Initialize next_column to 1
            next_column = 1

            # Find the next available column in the found row
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=cell.row, column=col).value is not None:
                    next_column = col + 1

            timestamp = s.starts_and_ends_of_stops[0]
            start_dt = datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H-%M-%S")
            # Write the new value to the next available column in the found row
            sheet.cell(row=cell.row, column=next_column, value=start_dt)  # training dt, in the first place of the array there is the start time
            sheet.cell(row=cell.row, column=next_column + 1, value=(float(s.number_of_repetitions_in_training) / s.max_repetitions_in_training))  # percent of the training that the patient managed to do
            sheet.cell(row=cell.row, column=next_column + 2, value=s.effort)  # effort rate the patient signed
            sheet.cell(row=cell.row, column=next_column + 3, value=calculate_training_length())  # percent of the training that the patient managed to do

            break  # Stop searching after finding the value

    workbook.save(file_path)


def which_welcome_record_to_say(headers_row=1):

    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_history_of_trainings"]

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if str(row[0].value) == str(s.chosen_patient_ID):  # Ensure comparison works for different types
            # Check if other cells in the row are not None or empty
            if any(cell.value for cell in row[1:]):  # Check if any other cell in the row has a value
                return "welcome"
            else:
                return "welcome_with_gymmy"

    # Return None if no matching patient is found
    return "welcome_with_gymmy"

# counts number of exercises in a training by ID by counting the true value
def count_number_of_exercises_in_training_by_ID():
    # Select the specific sheet
    workbook = openpyxl.load_workbook("Patients.xlsx")

    sheet = workbook["patients_exercises"]

    # Search for the row containing the search_value in the first column
    for row in sheet.iter_rows():
        if str(row[0].value) == s.chosen_patient_ID:
            row_number = row[0].row
            break
    else:
        # If the value is not found, return 0
        return 0

    # Get the values of the found row
    row_values = sheet[row_number]

    # Count the number of TRUE values in the row
    true_count = sum(1 for cell in row_values if cell.value is True)

    return true_count


def create_and_save_table_with_calculations(data, exercise):
    # Get session folder (created by create_workbook_for_training)
    session_folder = getattr(s, 'current_session_folder', None)
    
    if session_folder is None:
        # Fallback to old behavior if session_folder not set
        timestamp = s.starts_and_ends_of_stops[0]
        start_dt = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d_%H-%M-%S")
        session_folder = f"Patients/{s.chosen_patient_ID}/Trainings/{start_dt}"
    
    # Create exercise folder inside session folder (only when needed!)
    exercise_folder = f"{session_folder}/{exercise}"
    os.makedirs(exercise_folder, exist_ok=True)

    # Set the maximum title length
    max_title_length = 32  # As specified

    # Iterate over each table data (each table is for an angle/distance)
    for table_name, table_data in data.items():
        # Perform calculations (min, max, avg, std)
        # Create a new plot
        y_series = pd.Series(table_data['y'])
        y_values = y_series.dropna().tolist()

        # Center-pad the title to 32 characters
        title_text = table_name[:-2]
        display_title = title_text.center(max_title_length)  # Pad evenly on both sides

        if len(y_values) > 0:
            min_val = f"{min(y_values):.2f}"
            max_val = f"{max(y_values):.2f}"
            average = f"{(sum(y_values) / len(y_values)):.2f}"
            stdev = f"{np.std(y_values):.2f}"
        else:
            min_val = "אין נתונים"[::-1]
            max_val = "אין נתונים"[::-1]
            average = "אין נתונים"[::-1]
            stdev = "אין נתונים"[::-1]

        # Prepare data for the table
        calculation_data = {
            'ערכים'[::-1]: [min_val, max_val, average, stdev],  # Reverse Hebrew labels
            'מדדים'[::-1]: [s[::-1] for s in ['מינימום', 'מקסימום', 'ממוצע', 'סטיית תקן']]  # Reverse Hebrew labels
        }

        # Create a pandas DataFrame
        df = pd.DataFrame(calculation_data)

        # Create a new figure for the table only and set the background color
        fig, ax = plt.subplots(figsize=(2, 2))  # Adjust figure size to accommodate both table and header
        fig.patch.set_facecolor('#deeaf7')  # Set the background color of the figure

        # Hide axes completely (ensures no space around the table)
        ax.axis('off')

        # Add the table name as a header to the top of the figure
        ax.text(0.5, 0.9, display_title, ha='center', fontsize=13, weight='bold', transform=ax.transAxes)

        # Add the table to the figure with bold headers
        table = ax.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')

        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(12)
        table.scale(1.5, 1.3)  # Make the columns narrower by reducing the width scaling

        # Set the background color for the cells and the text properties
        for (i, j), cell in table.get_celld().items():
            if i == 0:  # Header row
                cell.set_text_props(weight='bold', fontsize=12)  # Set bold and increase font size
            else:
                cell.set_fontsize(12)  # Set a slightly smaller font for data rows

            cell.set_facecolor('#ffffff')  # White background for header cells

        # Save the table as an image with the background color and no transparency
        table_filename = f'{exercise_folder}/{table_name}.png'
        plt.savefig(table_filename, bbox_inches='tight', pad_inches=0, dpi=100)
        plt.close()  # Close the figure to clear memory
    
    print(f"[Excel] Saved tables to: {exercise_folder}")



def close_workbook():
    s.training_workbook.close()



def save_patient_rom(patient_id, rom_data):
    """
    Save patient ROM assessment results to Excel.
    
    Args:
        patient_id: Patient ID
        rom_data: Dict with format {'joint_name': angle_value, ...}
                  e.g., {'shoulder_flexion_right': 142.5, 'shoulder_flexion_left': 138.2}
    """
    print(f"\n{'='*60}")
    print(f"[ROM SAVE] Saving ROM data for patient {patient_id}")
    print(f"[ROM SAVE] Data to save:")
    for joint, value in rom_data.items():
        print(f"  - {joint}: {value:.2f}°")
    print(f"{'='*60}\n")
    
    """
    Save personalized ROM (Range of Motion) limits for a specific patient.
    Writes to a sheet named 'Patient_ROM' in Patients.xlsx.
    Creates the sheet if it doesn't exist.
    
    DEBUG VERSION - Comprehensive logging
    
    Args:
        patient_id: The ID of the patient to save ROM data for.
        rom_data: Dictionary of joint names to their max angle values.
                  Example: {'shoulder_flexion_right': 140, 'elbow_flexion_right': 110}
    
    Returns:
        True if ROM data was saved successfully, False otherwise.
    """
    import os
    
    file_path = "Patients.xlsx"
    sheet_name = "Patient_ROM"
    
    print("-" * 50)
    print("[Excel] save_patient_rom() called")
    print(f"[Excel]   patient_id: {patient_id}")
    print(f"[Excel]   rom_data: {rom_data}")
    print(f"[Excel]   file_path: {file_path}")
    print(f"[Excel]   sheet_name: {sheet_name}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"[Excel] ERROR: File {file_path} does not exist!")
        print(f"[Excel] Current working directory: {os.getcwd()}")
        print(f"[Excel] Files in current directory: {os.listdir('.')}")
        return False
    
    print(f"[Excel] File exists: OK")
    
    try:
        print("[Excel] Loading workbook...")
        workbook = openpyxl.load_workbook(file_path)
        print(f"[Excel] Workbook loaded. Sheets: {workbook.sheetnames}")
        
        # Create sheet if it doesn't exist
        if sheet_name not in workbook.sheetnames:
            print(f"[Excel] Sheet '{sheet_name}' not found. Creating it...")
            sheet = workbook.create_sheet(sheet_name)
            # Add patient_id as first header
            sheet.cell(row=1, column=1, value="patient_id")
            print(f"[Excel] Created new sheet '{sheet_name}' with 'patient_id' header")
        else:
            sheet = workbook[sheet_name]
            print(f"[Excel] Using existing sheet '{sheet_name}'")
        
        # Get existing headers (starting from column 1)
        print("[Excel] Reading existing headers...")
        existing_headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value is not None:
                existing_headers[cell.value] = col_idx
        print(f"[Excel] Existing headers: {existing_headers}")
        
        # Add any new joint names as headers
        next_col = sheet.max_column + 1 if sheet.max_column else 2
        print(f"[Excel] Next available column: {next_col}")
        
        for joint_name in rom_data.keys():
            if joint_name not in existing_headers:
                print(f"[Excel] Adding new header '{joint_name}' at column {next_col}")
                sheet.cell(row=1, column=next_col, value=joint_name)
                existing_headers[joint_name] = next_col
                next_col += 1
        
        # Find or create row for this patient
        print(f"[Excel] Looking for patient row with ID '{patient_id}'...")
        patient_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            cell_value = row[0].value
            if str(cell_value) == str(patient_id):
                patient_row = row[0].row
                print(f"[Excel] Found existing patient row: {patient_row}")
                break
        
        if patient_row is None:
            # Add new row for this patient
            patient_row = sheet.max_row + 1 if sheet.max_row else 2
            sheet.cell(row=patient_row, column=1, value=patient_id)
            print(f"[Excel] Created new patient row: {patient_row}")
        
        # Write ROM values
        print("[Excel] Writing ROM values...")
        for joint_name, value in rom_data.items():
            col_idx = existing_headers.get(joint_name)
            if col_idx:
                print(f"[Excel]   Writing {joint_name}={value} to row {patient_row}, column {col_idx}")
                sheet.cell(row=patient_row, column=col_idx, value=value)
            else:
                print(f"[Excel]   WARNING: Could not find column for '{joint_name}'")
        
        # Also update the timestamp for when ROM was last updated
        from datetime import datetime
        timestamp_col_name = "rom_last_updated"
        if timestamp_col_name not in existing_headers:
            timestamp_col = next_col
            sheet.cell(row=1, column=timestamp_col, value=timestamp_col_name)
            existing_headers[timestamp_col_name] = timestamp_col
            print(f"[Excel] Added '{timestamp_col_name}' column at {timestamp_col}")
        else:
            timestamp_col = existing_headers[timestamp_col_name]
        
        timestamp_value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=patient_row, column=timestamp_col, value=timestamp_value)
        print(f"[Excel]   Updated {timestamp_col_name} = {timestamp_value}")
        
        # Save the workbook
        print(f"[Excel] Saving workbook to {file_path}...")
        workbook.save(file_path)
        print("[Excel] Workbook saved successfully!")
        
        # Close the workbook to release the file handle
        workbook.close()
        print("[Excel] Workbook closed.")
        
        # Also update the global variable
        if hasattr(s, 'patient_rom_limits'):
            s.patient_rom_limits.update(rom_data)
            print(f"[Excel] Updated s.patient_rom_limits: {s.patient_rom_limits}")
        else:
            s.patient_rom_limits = rom_data.copy()
            print(f"[Excel] Created s.patient_rom_limits: {s.patient_rom_limits}")
        
        print(f"[Excel] SUCCESS: Saved ROM limits for patient {patient_id}")
        print("-" * 50)
        return True
        
    except PermissionError:
        print(f"[Excel] ERROR: Permission denied! Is the file open in Excel?")
        print(f"[Excel] Close the file and try again.")
        return False
    except FileNotFoundError:
        print(f"[Excel] ERROR: File {file_path} not found!")
        return False
    except Exception as e:
        print(f"[Excel] ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False



def load_patient_rom(patient_id):
    """
    Load personalized ROM (Range of Motion) limits for a specific patient.
    Reads from a sheet named 'Patient_ROM' in Patients.xlsx and populates s.patient_rom_limits.
    
    DEBUG VERSION - Comprehensive logging
    
    Args:
        patient_id: The ID of the patient to load ROM data for.
    
    Returns:
        True if ROM data was loaded successfully, False otherwise.
    """
    file_path = "Patients.xlsx"
    sheet_name = "Patient_ROM"
    
    print("-" * 50)
    print("[Excel] load_patient_rom() called")
    print(f"[Excel]   patient_id: {patient_id}")
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"[Excel] Workbook loaded. Sheets: {workbook.sheetnames}")
        
        # Check if the Patient_ROM sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"[Excel] Sheet '{sheet_name}' not found. No ROM data available.")
            s.patient_rom_limits = {}
            workbook.close()
            return False
        
        sheet = workbook[sheet_name]
        print(f"[Excel] Using sheet '{sheet_name}'")
        
        # Find the row for this patient (patient_id in first column)
        patient_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            if str(row[0].value) == str(patient_id):
                patient_row = row[0].row
                break
        
        if patient_row is None:
            print(f"[Excel] Patient {patient_id} not found in {sheet_name}.")
            s.patient_rom_limits = {}
            workbook.close()
            return False
        
        print(f"[Excel] Found patient at row {patient_row}")
        
        # Read header row to get joint names
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(cell.value)
        print(f"[Excel] Headers: {headers}")
        
        # Populate patient_rom_limits dictionary
        s.patient_rom_limits = {}
        for col_idx, header in enumerate(headers[1:], start=2):  # Skip first column (patient_id)
            cell_value = sheet.cell(row=patient_row, column=col_idx).value
            if cell_value is not None and header is not None:
                try:
                    s.patient_rom_limits[header] = float(cell_value)
                except (ValueError, TypeError):
                    pass  # Skip non-numeric values
        
        workbook.close()
        
        print(f"[Excel] Loaded ROM limits: {s.patient_rom_limits}")
        print("-" * 50)
        return True
        
    except FileNotFoundError:
        print(f"[Excel] ERROR: File {file_path} not found!")
        s.patient_rom_limits = {}
        return False
    except Exception as e:
        print(f"[Excel] ERROR: {type(e).__name__}: {e}")
        s.patient_rom_limits = {}
        return False














# def create_summary_workbook():
#         workbook_name = f"Patients/{s.chosen_patient_ID}/summary.xlsx"
#
#         # Create a new workbook
#         workbook = xlsxwriter.Workbook(workbook_name)
#
#         # Add a worksheet
#         workbook.add_worksheet("Sheet1")
#
#         # Close the workbook to save it
#         workbook.close()


# def add_exercise_to_summary(exercise_name, avg, sd, min_val, max_val, time_val):
#     # Load the existing workbook
#     file_path = f"Patients/{s.chosen_patient_ID}/summary.xlsx"
#     workbook = openpyxl.load_workbook(file_path)
#
#     # Check if "Sheet1" exists and rename it to the exercise name
#     if "Sheet1" in workbook.sheetnames:
#         sheet = workbook["Sheet1"]
#         sheet.title = exercise_name
#     # Check if the exercise sheet already exists, if not, create it
#     elif exercise_name not in workbook.sheetnames:
#         sheet = workbook.create_sheet(title=exercise_name)
#
#     # Add headers if it's a new or renamed sheet
#     if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:  # Check if it's an empty sheet
#         sheet.append(["Date Time", "Average", "Standard Deviation", "Minimum", "Maximum", "Time"])
#
#     # Find the next empty row (last row + 1)
#     next_row = sheet.max_row + 1
#
#     # Add data to the next available row, with the first column being the datetime
#     sheet.cell(row=next_row, column=1, value=s.start_dt)  # Datetime from s.start_dt
#     sheet.cell(row=next_row, column=2, value=avg)  # Average
#     sheet.cell(row=next_row, column=3, value=sd)  # Standard Deviation
#     sheet.cell(row=next_row, column=4, value=min_val)  # Minimum
#     sheet.cell(row=next_row, column=5, value=max_val)  # Maximum
#     sheet.cell(row=next_row, column=6, value=time_val)  # Time
#
#     # Save the workbook
#     workbook.save(file_path)









if __name__ == "__main__":

    s.chosen_patient_ID="55581599"
    print(which_welcome_record_to_say())
    # Example usage
